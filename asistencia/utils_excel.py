import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone


def generar_reporte_excel(queryset, titulo_periodo, filename_prefix="reporte"):
    """Genera un HttpResponse con un Excel con formato profesional a partir
    de un queryset de RegistroAsistencia."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Ingreso"

    headers = ['Apellidos', 'Nombres', 'DNI', 'Fecha', 'Hora Entrada', 'Hora Salida', 'Horas Trabajadas', 'Estado']

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    ws.append(['CONTROL DE INGRESO Y SALIDA DE PERSONAL'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws['A1'].alignment = center

    ws.append([titulo_periodo])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws['A2'].font = Font(italic=True, size=10, color="333333")
    ws['A2'].alignment = center

    ws.append([f"Generado el: {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M:%S')}"])
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    ws['A3'].font = Font(italic=True, size=9, color="666666")
    ws['A3'].alignment = center

    ws.append([])

    header_row_idx = 5
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    row_idx = header_row_idx + 1
    for registro in queryset.select_related('trabajador').order_by('-fecha', 'trabajador__apellidos'):
        hora_entrada = timezone.localtime(registro.hora_entrada).strftime('%H:%M:%S') if registro.hora_entrada else '-'
        hora_salida = timezone.localtime(registro.hora_salida).strftime('%H:%M:%S') if registro.hora_salida else '-'
        fila = [
            registro.trabajador.apellidos,
            registro.trabajador.nombres,
            registro.trabajador.dni or '-',
            registro.fecha.strftime('%d/%m/%Y'),
            hora_entrada,
            hora_salida,
            registro.horas_trabajadas,
            registro.estado,
        ]
        ws.append(fila)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.alignment = center
            if registro.estado == 'EN CURSO':
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row_idx += 1

    for col_num, header in enumerate(headers, start=1):
        max_length = len(header)
        for row in range(header_row_idx + 1, row_idx):
            value = ws.cell(row=row, column=col_num).value
            if value:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 4

    ws.freeze_panes = f"A{header_row_idx + 1}"
    ws.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{header_row_idx}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{filename_prefix}_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
